import openpyxl

#load the xl file to inv_file and take the sheet with the items to product_list
inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

#dictionary
products_per_supplier = {}
total_value_per_supplier = {}
prod_under_ten_inv = {}

#go threw all item rows
for product_row in range(2,product_list.max_row + 1):
    #access the value of a cell
    supplier_name = product_list.cell(product_row,4).value
    inventory = product_list.cell(product_row,2).value
    price=product_list.cell(product_row,3).value
    prod_num= product_list.cell(product_row,1).value
    inventory_price = product_list.cell(product_row, 5)

    #calculate the items for each supplier
    if supplier_name in products_per_supplier:
        #total items
        products_per_supplier[supplier_name] += 1
        #total value
        total_value_per_supplier[supplier_name] += inventory * price
    else:
        #first introduction for supplier
        products_per_supplier[supplier_name] = 1
        total_value_per_supplier[supplier_name] = inventory * price

    #take items with inv less then 10
    if inventory<10:
        prod_under_ten_inv[prod_num] = inventory

    #add value for total inventory price
    inventory_price.value= inventory * price

print(products_per_supplier)
print(total_value_per_supplier)
print(prod_under_ten_inv)

#save changes to new file
inv_file.save("inventory_with_total_value.xlsx")